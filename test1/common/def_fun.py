# TIME: 2021/4/22 11:30
# author: xiaoyu

import requests
import openpyxl

#读取测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 打开工作簿
    sh = wb[sheetname]  # 获取sheet, 获取表单
    rows = sh.max_row  # 获取register表里面最大的行号，自定义一个变量rows来接收数据
    case_list = []  # 创建一个空的列表，用来存储测试用例

    for i in range(2, rows + 1):
        dict1 = dict(
            case_id=sh.cell(row=i, column=1).value,  # key = value
            url=sh.cell(row=i, column=5).value,  # key = value
            data=sh.cell(row=i, column=6).value,  # key = value
            expected=sh.cell(row=i, column=7).value  # key = value
        )
        case_list.append(dict1)  # 3. 接下来要把字典里面的值，循环的放到1个空list里面
    return case_list  # 定义函数需要有个返回值，给别人调用时使用

#接口请求
def api_ful(url,data):  #定义函数，参数
    # url_login = 'http://8.129.91.152:8766/futureloan/member/login'
    # data_login = {"mobile_phone": "18720031763", "pwd": "lemon666"}
    headers= {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}

    result = requests.post(url=url, json=data, headers=headers).json()   #接收响应正文的结果
    return result #返回结果

#（写入excel结果函数）
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)  #打开工作簿
    sh = wb[sheetname] #获取sheet, 获取表单
    sh.cell(row=row,column=column).value=final_result   #因为每一行的第8列都需要写入结果，写入的结果也是会变的，所以也要作为参数final_result
    wb.save(filename)   #保存excel


#接口测试实战
#调用读取用例函数，调用接口请求函数（登录、注册），调用写入结果的函数  ---封装成函数
def excute(filename,sheetname,):
    cases = read_data(filename,sheetname)
    for dict1 in cases:
        # print(dict1)
        # print(type(i))
        case_id =dict1['case_id']
        url = dict1['url']
        data = eval(dict1['data'])
        # print(type(data))
        # print(case_id,url,data,expected)

        result_api = api_ful(url=url,data =data)
        print(result_api)

        expected = eval(dict1['expected'])  #获取预期结果的code和msg
        code = expected['code']
        msg = expected['msg']
        print("预期结果code: {},msg :{}".format(code,msg))

        real_code = result_api['code']
        real_msg = result_api['msg']

        print("实际结果code: {},msg :{}".format(real_code,real_msg))

        if code == real_code and msg==real_msg:
            print('第{}用例测试通过'.format(case_id))
            finl_result = 'passed'
        else:
            print('第{}用例测试不通过'.format(case_id))
            finl_result = 'false'

        write_result(filename,sheetname,case_id+1,8,finl_result)

        print('-' * 100)

# excute('C:\\tool\\python_workspacea\\test1\\testdata\\test_case_api.xlsx','register')
excute('C:\\tool\\python_workspacea\\test1\\testdata\\test_case_api.xlsx','login')

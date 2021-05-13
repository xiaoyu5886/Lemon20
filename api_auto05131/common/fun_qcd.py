# TIME: 2021/5/13 9:46
# author: xiaoyu

from openpyxl import load_workbook
import requests

# log_url ="http://8.129.91.152:8766/futureloan/member/login"
# log_data ={"mobile_phone":"15512345678","pwd":"12345678"}
# "C:\\tool\python_workspacea\\api_auto16\\test_data\\test_case_api.xlsx"
# log_headers ={'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}

#返回响应结果
def api_log(url,data):
    # url="http://8.129.91.152:8766/futureloan/member/login"
    # data ={"mobile_phone":"15512345678","pwd":"12345678"}
    headers ={'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
    log_res = requests.post(url=url,json=data,headers=headers).json()
    # print(log_res)
    return log_res

# 读取测试用例
def read_case(filename,sheetname):
    wb =load_workbook(filename)
    sh =wb[sheetname]
    row =sh.max_row
    # print(row)
    case_list =[]
    for i in range(2,row+1):
        dict1 =dict(id = sh.cell(i,column=1).value,url = sh.cell(i,column=5).value,data = sh.cell(i,column=6).value,exceped = sh.cell(i,column=7).value)
        # print(dict1)
        case_list.append(dict1)
    return case_list

# 写入测试结果到excel
def write_case(filename,sheetname,row,column,final_write):
    wb =load_workbook(filename)
    sh =wb[sheetname]
    sh.cell(row=row,column=column).value = final_write
    wb.save(filename)
# 对响应结果和测试用例的断言进行比较，然后写入结果

